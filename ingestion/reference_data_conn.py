"""
Reference Data connector.

Ingests the SWP EOD Data Feeds Reference List — a flat field-reference catalog:
one row per field, with category, position, field name, short description, and
detailed description. Populates the reference_data table.

Each field name is normalized with the SAME logic as the datapoint indexer, so
reference rows attach to the right data point. Attachment key is
(category + normalized field name) — the same field can carry a different
description under different categories. Rows whose normalized field name is not
found in dp_registry are flagged resolved='N' (a reference/catalog gap).

Env: REFERENCE_DATA_XLSX = path to the reference workbook (flat, first sheet read).
Run AFTER datapoint_index (so dp_registry exists to resolve against) and BEFORE
search_index.
"""
from __future__ import annotations
import os
import re
import logging

log = logging.getLogger("cp.reference_data")


def _normalize(name: str) -> str:
    """IDENTICAL to datapoint_indexer._normalize so keys match."""
    s = name or ""
    if re.search(r"[a-z][A-Z]", s):
        s = re.sub(r"(?<=[a-z])(?=[A-Z])", "_", s)
    s = re.sub(r"[^a-zA-Z0-9]+", "_", s).strip("_").lower()
    return s


def _norm_header(h):
    return "".join(ch for ch in str(h or "").lower() if ch.isalnum())


class ReferenceDataConnector:
    """Step name: reference_data."""

    def __init__(self, path=None):
        self.path = path or os.environ.get("REFERENCE_DATA_XLSX")

    @classmethod
    def from_env(cls, resolver=None):
        return cls()

    def parse(self):
        if not self.path or not os.path.exists(self.path):
            log.warning("reference_data: REFERENCE_DATA_XLSX not set or missing (%s); skipping", self.path)
            return None
        from openpyxl import load_workbook
        wb = load_workbook(self.path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]   # flat sheet, first one
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return []
        header = [_norm_header(h) for h in rows[0]]
        idx = {h: i for i, h in enumerate(header)}

        def cell(row, *names):
            for n in names:
                k = _norm_header(n)
                if k in idx and idx[k] < len(row):
                    v = row[idx[k]]
                    if v is not None and str(v).strip() != "":
                        return str(v).strip()
            return None

        out = []
        last_category = None   # forward-fill: category may appear once atop a block
        for row in rows[1:]:
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            fname = cell(row, "Field Name", "Field", "Name", "Attribute")
            if not fname:
                continue
            pos = cell(row, "Position", "Pos", "Order", "Seq")
            cat = cell(row, "Category", "Group", "Section", "Workstream", "Module", "Entity")
            if cat:
                last_category = cat
            else:
                cat = last_category   # inherit from the block header above
            out.append({
                "category": cat,
                "position": int(float(pos)) if pos and str(pos).replace(".", "").isdigit() else None,
                "field_name": fname,
                "field_description": cell(row, "Field Description", "Description", "Short Description", "Business Meaning"),
                "detail_description": cell(row, "Detail Description", "Detailed Description", "Detail", "Details", "Long Description", "Notes"),
            })
        log.info("reference_data: parsed %d reference rows", len(out))
        return out

    def load(self, loader, bundle=None):
        rows = self.parse() if bundle is None else bundle
        if not rows:
            return
        conn = loader.conn
        cur = conn.cursor()

        known_dps = set()
        try:
            cur.execute("SELECT dp_name_normalized FROM dp_registry")
            known_dps = {(r[0] or "").lower() for r in cur.fetchall()}
        except Exception as e:
            log.warning("reference_data: could not read dp_registry (%s); resolution all N", e)

        unresolved = []
        seen = set()
        for r in rows:
            norm = _normalize(r["field_name"])
            cat = r.get("category") or ""
            ref_id = f"{cat}|{norm}"[:400]
            if ref_id in seen:   # avoid dup PK within one run
                continue
            seen.add(ref_id)
            resolved = "Y" if norm in known_dps else "N"
            if resolved == "N":
                unresolved.append((cat, r["field_name"]))
            loader._merge("reference_data", ("ref_id",), {
                "ref_id": ref_id, "category": cat or None, "position_order": r.get("position"),
                "field_name": r["field_name"][:256],
                "field_name_normalized": norm[:256],
                "field_description": (r.get("field_description") or "")[:2000] or None,
                "detail_description": r.get("detail_description"),
                "resolved": resolved, "project_id": "sei",
            })
        loader.commit()

        if unresolved:
            log.warning("reference_data: %d reference fields did NOT resolve to dp_registry:", len(unresolved))
            for cat, fn in unresolved[:25]:
                log.warning("   unresolved '%s' in category '%s'", fn, cat)
            if len(unresolved) > 25:
                log.warning("   ... +%d more", len(unresolved) - 25)
        log.info("reference_data: loaded %d rows; %d unresolved", len(seen), len(unresolved))
        return {"loaded": len(seen), "unresolved": unresolved}
